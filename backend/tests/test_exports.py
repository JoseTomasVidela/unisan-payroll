from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from sqlalchemy import select

from app.models import Employee, PayrollExportLog, PayrollManualAdjustment
from app.softland import ensure_softland_mappings
from conftest import login
from test_settlements import seed_settlement


def admin_headers(client):
    token = login(client, "admin", "admin-password")
    return {"Authorization": f"Bearer {token}"}


def user_headers(client):
    token = login(client, "consulta", "consulta-password")
    return {"Authorization": f"Bearer {token}"}


def test_export_individual_liquidation_excel(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        (
            f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}"
            "&cost_center=DR&role_type=DRIVER&file_format=xlsx"
        ),
        headers=admin_headers(client),
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    assert 'filename="Chofer Uno-Junio 2026.xlsx"' in response.headers["content-disposition"]

    archive = ZipFile(BytesIO(response.content))
    worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "TRABAJADOR: Chofer Uno" in worksheet_xml
    assert "Ciclo</t>" not in worksheet_xml
    assert "Centro</t>" not in worksheet_xml
    assert "Vista</t>" not in worksheet_xml
    assert "Actividad" in worksheet_xml
    assert "Tarifa" in worksheet_xml
    assert "Evento" in worksheet_xml
    assert "TOTAL A PAGAR" in worksheet_xml
    assert "PRODUCCION TOTAL" in worksheet_xml


def test_export_visible_liquidations_in_one_excel_sheet(client, db_factory):
    driver_id, assistant_id = seed_settlement(db_factory)
    response = client.get(
        f"/api/exports/settlements.xlsx?items=1:{driver_id},1:{assistant_id}",
        headers=admin_headers(client),
    )
    assert response.status_code == 200
    assert 'filename="Liquidaciones seleccionadas.xlsx"' in response.headers["content-disposition"]
    archive = ZipFile(BytesIO(response.content))
    worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
    assert "TRABAJADOR: Chofer Uno" in worksheet_xml
    assert "TRABAJADOR: Auxiliar Uno" in worksheet_xml
    assert worksheet_xml.index("TRABAJADOR: Chofer Uno") < worksheet_xml.index("TRABAJADOR: Auxiliar Uno")
    assert workbook_xml.count("<sheet ") == 1
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    title_rows = [
        row
        for row in range(1, worksheet.max_row + 1)
        if str(worksheet.cell(row, 1).value or "").startswith("TRABAJADOR:")
    ]
    assert len(title_rows) == 2
    assert worksheet.cell(title_rows[0], 1).fill.fgColor.rgb == "0029405A"
    assert worksheet.cell(title_rows[0], 1).border.left.style == "medium"
    assert worksheet.cell(title_rows[1], 1).border.left.style == "medium"


def test_export_individual_liquidation_csv(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        (
            f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}"
            "&cost_center=DR&role_type=DRIVER&file_format=csv"
        ),
        headers=admin_headers(client),
    )

    assert response.status_code == 200
    body = response.content.decode("utf-8-sig")
    assert "Trabajador,Chofer Uno,,," in body
    assert "Actividad,Unidades,Tarifa,Total" in body
    assert "Evento,5,5,25" in body
    assert "TOTAL A PAGAR,,," in body
    assert "PRODUCCION TOTAL,,," in body


def test_export_individual_liquidation_pdf(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)
    with db_factory() as db:
        db.add(
            PayrollManualAdjustment(
                cycle_id=1,
                employee_id=driver_id,
                cost_center="ALL",
                role_type="ALL",
                adjustment_type="BONUS",
                adjustment_name="Bono exportado",
                units=Decimal("3"),
                amount=Decimal("2000"),
            )
        )
        db.add(
            PayrollManualAdjustment(
                cycle_id=1,
                employee_id=driver_id,
                cost_center="ALL",
                role_type="ALL",
                adjustment_type="EVENT_BONUS",
                adjustment_name="Bono Evento",
                units=Decimal("1"),
                amount=Decimal("20"),
            )
        )
        db.commit()

    response = client.get(
        (
            f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}"
            "&cost_center=DR&role_type=DRIVER&file_format=pdf"
        ),
        headers=admin_headers(client),
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/pdf")
    assert response.content.startswith(b"%PDF-")
    assert 'filename="Chofer Uno-Junio 2026.pdf"' in response.headers["content-disposition"]
    assert b"MOVILES DE CHILE S.A." in response.content
    assert b"ANEXO DE LIQUIDACION DE SUELDO MES DE JUNIO 2026" in response.content
    assert b"Articulo 54 bis inciso tercero del Codigo del Trabajo" in response.content
    assert b"Chofer Uno" in response.content
    assert b"DESCRIPCION" in response.content
    assert b"A PAGAR" in response.content
    assert b"SEMANA CORRIDA" in response.content
    assert b"UNIBOX" in response.content
    assert b"BONO EXPORTADO" in response.content
    assert b"6.000" in response.content
    assert b"TOTAL" in response.content
    assert b"FIRMA TRABAJADOR" in response.content


def test_pdf_always_displays_unibox_even_without_matching_activity():
    from app.exporter import export_pdf_bytes

    pdf = export_pdf_bytes({
        "employee": {"employee_name": "Sin Unibox", "rut": ""},
        "cycle": {"cycle_name": "Ciclo Junio 2026"},
        "rows": [{
            "row_type": "adjustment_production_bonus",
            "concept_code": "ADJUSTMENT_PRODUCTION_BONUS",
            "concept_name": "Bono Producción - prueba texto",
            "units": Decimal("2"),
            "rate": Decimal("2000"),
            "total": Decimal("4000"),
        }],
        "total_to_pay": Decimal("0"),
        "production_total": Decimal("4000"),
    })

    assert b"UNIBOX" in pdf


def test_export_from_search_uses_consolidated_scope_and_logs(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}&file_format=csv",
        headers=admin_headers(client),
    )

    assert response.status_code == 200

    with db_factory() as db:
        logs = db.scalars(select(PayrollExportLog).order_by(PayrollExportLog.id)).all()
        assert len(logs) == 1
        assert logs[0].export_scope == "SEARCH"
        assert logs[0].file_format == "CSV"
        assert logs[0].cycle_id == 1
        assert logs[0].employee_id == driver_id
        assert logs[0].cost_center is None
        assert logs[0].role_type is None


def test_export_logs_context_export(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        (
            f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}"
            "&cost_center=DR&role_type=DRIVER&file_format=csv"
        ),
        headers=admin_headers(client),
    )

    assert response.status_code == 200

    with db_factory() as db:
        log = db.scalar(select(PayrollExportLog).where(PayrollExportLog.employee_id == driver_id))
        assert log is not None
        assert log.export_scope == "SETTLEMENT"
        assert log.file_format == "CSV"
        assert log.cost_center == "DR"
        assert log.role_type == "DRIVER"
        assert log.user_id == 1
        assert log.file_name.endswith(".csv")


def test_authenticated_user_can_export(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}&file_format=csv",
        headers=user_headers(client),
    )

    assert response.status_code == 200
    assert "Actividad,Unidades,Tarifa,Total" in response.content.decode("utf-8-sig")


def test_export_requires_authentication(client, db_factory):
    driver_id, _ = seed_settlement(db_factory)

    response = client.get(
        f"/api/exports/settlement?cycle_id=1&employee_id={driver_id}&file_format=csv"
    )

    assert response.status_code == 401


def test_export_softland_cycle_has_exact_columns_order_and_aggregates(client, db_factory):
    driver_id, assistant_id = seed_settlement(db_factory)
    with db_factory() as db:
        db.get(Employee, driver_id).rut = "12.584.663-4"
        db.get(Employee, assistant_id).rut = "9.876.543-2"
        db.add_all(
            [
                PayrollManualAdjustment(
                    cycle_id=1,
                    employee_id=driver_id,
                    cost_center="ALL",
                    role_type="ALL",
                    adjustment_type="BONUS",
                    adjustment_name="Bono",
                    units=Decimal("1"),
                    amount=Decimal("100"),
                ),
                PayrollManualAdjustment(
                    cycle_id=1,
                    employee_id=driver_id,
                    cost_center="ALL",
                    role_type="ALL",
                    adjustment_type="EVENT_BONUS",
                    adjustment_name="Bono Evento",
                    units=Decimal("1"),
                    amount=Decimal("20"),
                ),
                PayrollManualAdjustment(
                    cycle_id=1,
                    employee_id=driver_id,
                    cost_center="ALL",
                    role_type="ALL",
                    adjustment_type="VACATION",
                    adjustment_name="Vacaciones",
                    units=Decimal("1"),
                    amount=Decimal("30"),
                ),
                PayrollManualAdjustment(
                    cycle_id=1,
                    employee_id=driver_id,
                    cost_center="ALL",
                    role_type="ALL",
                    adjustment_type="VACATION_BONUS",
                    adjustment_name="Bono Vacaciones",
                    units=Decimal("1"),
                    amount=Decimal("40"),
                ),
            ]
        )
        db.commit()
        ensure_softland_mappings(db)

    response = client.get(
        "/api/exports/softland?cycle_id=1",
        headers=admin_headers(client),
    )

    assert response.status_code == 200
    assert 'filename="Export Softland (06-2026).xlsx"' in response.headers["content-disposition"]
    archive = ZipFile(BytesIO(response.content))
    worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "FICHA" in worksheet_xml
    assert "CODI" in worksheet_xml
    assert "MES AÑO" in worksheet_xml
    assert "VALOR" in worksheet_xml
    assert worksheet_xml.index("H005") < worksheet_xml.index("H008")
    assert worksheet_xml.index("H008") < worksheet_xml.index("H022")
    assert worksheet_xml.index("H022") < worksheet_xml.index("H040")
    assert "12584663" in worksheet_xml
    assert "9876543" not in worksheet_xml
    assert "06/2026" in worksheet_xml
    assert "<v>150</v>" in worksheet_xml  # H008: 50 de producción + 100 de bono.
    assert "<v>45</v>" in worksheet_xml  # H022: 25 de eventos + 20 de bono evento.
    assert "<v>70</v>" in worksheet_xml  # H040: vacaciones + bono vacaciones.
    assert "<v>0</v>" not in worksheet_xml

    with db_factory() as db:
        logs = db.scalars(
            select(PayrollExportLog).where(PayrollExportLog.export_scope == "SOFTLAND")
        ).all()
        assert len(logs) == 1


def test_export_softland_uses_worker_name_when_rut_is_missing(client, db_factory):
    seed_settlement(db_factory)
    with db_factory() as db:
        ensure_softland_mappings(db)

    response = client.get(
        "/api/exports/softland?cycle_id=1",
        headers=admin_headers(client),
    )

    assert response.status_code == 200
    archive = ZipFile(BytesIO(response.content))
    worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "Chofer Uno" in worksheet_xml
    assert "Auxiliar Uno" not in worksheet_xml
    assert "H040" not in worksheet_xml
    assert "<v>0</v>" not in worksheet_xml


def test_export_softland_reports_unmapped_adjustment(client, db_factory):
    driver_id, assistant_id = seed_settlement(db_factory)
    with db_factory() as db:
        db.get(Employee, driver_id).rut = "12.584.663-4"
        db.get(Employee, assistant_id).rut = "9.876.543-2"
        db.add(
            PayrollManualAdjustment(
                cycle_id=1,
                employee_id=driver_id,
                cost_center="ALL",
                role_type="ALL",
                adjustment_type="MANUAL_ADJUSTMENT",
                adjustment_name="Ajuste sin código",
                units=Decimal("1"),
                amount=Decimal("100"),
            )
        )
        db.commit()
        ensure_softland_mappings(db)

    response = client.get(
        "/api/exports/softland?cycle_id=1",
        headers=admin_headers(client),
    )

    assert response.status_code == 422
    assert "Conceptos sin homologación Softland" in response.json()["detail"]
    assert "Ajuste sin código" in response.json()["detail"]
