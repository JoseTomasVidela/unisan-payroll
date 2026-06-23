from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.datetime import from_excel
from sqlalchemy import select
from sqlalchemy.orm import Session

from .employee_names import names_refer_to_same_person
from .models import Employee, PayrollCycle, PayrollImport, PayrollRecord, User

INVALID_PERSON_VALUES = {"", "0", "n/a", "sin auxiliar"}
ZERO = Decimal("0")

DR_HEADERS = [
    "Operador", "Auxiliar", "Cod. Usuario", "Fecha Inicial", "Fecha Final",
    "Duración", "Estado", "Despacho / Retiro", "Entrada < 19:30",
    "Salida > 19:30", "Feria Semana 01", "Feria Semana 02",
    "Fuera Radio Normal", "Fuera Radio V Región", "Sabado", "Domingo",
    "Viajes Por Cliente", "Sábado > 16:00", "Domingo > 16:00", "URL",
    "Secado", "Evento", "Punto de Agua", "Basurero Grande",
    "Basurero chico", "Fosa", "Aux Sabado Semana 02",
    "Aux Domingo Semana 02",
]

SERVICES_HEADERS = [
    "Operador", "Codigo de Usuario", "Auxiliar1", "Auxiliar2",
    "Fecha Inicial", "Fecha Final", "Duracion", "Estado", "Aseo", "Secado",
    "Despacho / Retiro", "Entrada < 07:30", "Salida > 19:30",
    "Fuera Radio Normal", "Fuera Radio V Región", "Entrega Kit",
    "Carga Lavamanos", "Sabado", "Domingo", "Aseo Fin de Semana",
    "Secado Fin de Semana", "Sabado > 16:00", "Domingo > 16:00",
    "Succión Riles (M3)", "URL", None,
]

COMMON_ZERO_FIELDS = {
    "dispatch_flag", "entry_before_1930_qty", "entry_before_0730_qty",
    "exit_after_1930_qty", "fair_week_1_flag", "fair_week_2_flag",
    "outside_radius_flag", "outside_radius_v_region_qty", "saturday_week_1_qty",
    "sunday_week_1_qty", "saturday_week_2_qty", "sunday_week_2_qty",
    "client_trips_qty", "saturday_after_1600_qty", "sunday_after_1600_qty",
    "cleaning_flag", "drying_flag", "weekend_cleaning_qty",
    "weekend_drying_qty", "event_flag", "water_point_flag",
    "kit_delivery_flag", "lavatory_load_flag", "large_trash_bin_qty",
    "small_trash_bin_qty", "fosa_qty", "riles_suction_flag",
}

DR_CONCEPTS = {
    7: "dispatch_flag",
    8: "entry_before_1930_qty",
    9: "exit_after_1930_qty",
    10: "fair_week_1_flag",
    11: "fair_week_2_flag",
    12: "outside_radius_flag",
    13: "outside_radius_v_region_qty",
    14: "saturday_week_1_qty",
    15: "sunday_week_1_qty",
    16: "client_trips_qty",
    17: "saturday_after_1600_qty",
    18: "sunday_after_1600_qty",
    20: "weekend_drying_qty",
    21: "event_flag",
    22: "water_point_flag",
    23: "large_trash_bin_qty",
    24: "small_trash_bin_qty",
    25: "fosa_qty",
}

SERVICES_CONCEPTS = {
    8: "cleaning_flag",
    9: "drying_flag",
    10: "dispatch_flag",
    11: "entry_before_0730_qty",
    12: "exit_after_1930_qty",
    13: "outside_radius_flag",
    14: "outside_radius_v_region_qty",
    15: "kit_delivery_flag",
    16: "lavatory_load_flag",
    17: "saturday_week_1_qty",
    18: "sunday_week_1_qty",
    19: "weekend_cleaning_qty",
    20: "weekend_drying_qty",
    21: "saturday_after_1600_qty",
    22: "sunday_after_1600_qty",
    23: "riles_suction_flag",
}


@dataclass
class RecordCandidate:
    values: dict[str, Any]

    @property
    def duplicate_key(self) -> tuple[int, str, str, str]:
        return (
            self.values["cycle_id"],
            self.values["role_type"],
            self.values["source_person_slot"],
            self.values["source_row_hash"],
        )


@dataclass
class ParsedImport:
    source_type: str
    cost_center: str
    rows_read: int = 0
    rows_outside_cycle: int = 0
    candidates: list[RecordCandidate] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class CycleDefinition:
    cycle_name: str
    start_date: date
    end_date: date


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def is_valid_person(value: Any) -> bool:
    return normalize_text(value).casefold() not in INVALID_PERSON_VALUES


def to_decimal(value: Any) -> Decimal:
    if value is None or isinstance(value, bool):
        return ZERO
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = normalize_text(value).replace(" ", "")
    if not text:
        return ZERO
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return ZERO


def to_work_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_duration_minutes(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 60)
    if isinstance(value, time):
        return value.hour * 60 + value.minute + round(value.second / 60)
    if isinstance(value, (int, float, Decimal)):
        return round(float(value) * 24 * 60)
    text = normalize_text(value)
    parts = text.split(":")
    if len(parts) in (2, 3) and all(part.isdigit() for part in parts):
        hours, minutes = int(parts[0]), int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
        return hours * 60 + minutes + round(seconds / 60)
    return None


def canonical_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value.total_seconds())
    if isinstance(value, (int, float, Decimal)):
        return format(Decimal(str(value)).normalize(), "f")
    return normalize_text(value)


def row_hash(values: tuple[Any, ...], relevant_indexes: list[int]) -> str:
    payload = [canonical_value(values[index]) for index in relevant_indexes]
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def validate_headers(actual: tuple[Any, ...], expected: list[str | None]) -> None:
    normalized = [normalize_text(item) or None for item in actual[: len(expected)]]
    if normalized != expected:
        differences = [
            f"{index + 1}: esperado={wanted!r}, recibido={received!r}"
            for index, (wanted, received) in enumerate(zip(expected, normalized))
            if wanted != received
        ]
        raise ValueError("Encabezados Excel inválidos. " + "; ".join(differences[:8]))


def load_source_sheet(content: bytes, source_type: str):
    if source_type not in {"DR", "SERVICES"}:
        raise ValueError("source_type debe ser DR o SERVICES.")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValueError("El archivo seleccionado no es un Excel .xlsx válido.") from exc
    if "Base Datos" not in workbook.sheetnames:
        raise ValueError("El archivo no contiene la hoja obligatoria 'Base Datos'.")
    sheet = workbook["Base Datos"]
    expected_headers = DR_HEADERS if source_type == "DR" else SERVICES_HEADERS
    header_row = next(sheet.iter_rows(max_row=1), None)
    if header_row is None:
        raise ValueError("La hoja 'Base Datos' está vacía.")
    validate_headers(tuple(cell.value for cell in header_row), expected_headers)
    return workbook, sheet, expected_headers


SPANISH_MONTHS = (
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)


def cycle_definition_for_date(work_date: date) -> CycleDefinition:
    if work_date.day >= 22:
        start_date = date(work_date.year, work_date.month, 22)
        if work_date.month == 12:
            closing_year, closing_month = work_date.year + 1, 1
        else:
            closing_year, closing_month = work_date.year, work_date.month + 1
    else:
        closing_year, closing_month = work_date.year, work_date.month
        if work_date.month == 1:
            start_date = date(work_date.year - 1, 12, 22)
        else:
            start_date = date(work_date.year, work_date.month - 1, 22)
    end_date = date(closing_year, closing_month, 21)
    return CycleDefinition(
        cycle_name=f"Ciclo {SPANISH_MONTHS[closing_month]} {closing_year}",
        start_date=start_date,
        end_date=end_date,
    )


def ensure_workbook_cycles(
    db: Session,
    content: bytes,
    source_type: str,
) -> tuple[dict[tuple[date, date], PayrollCycle], int]:
    workbook, sheet, expected_headers = load_source_sheet(content, source_type)
    date_index = 3 if source_type == "DR" else 4
    definitions: dict[tuple[date, date], CycleDefinition] = {}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(row[: len(expected_headers)])
        if not any(value not in (None, "") for value in values):
            continue
        work_date = to_work_date(values[date_index], workbook.epoch)
        if work_date is None:
            raise ValueError(f"Fila {row_number}: Fecha Inicial inválida.")
        definition = cycle_definition_for_date(work_date)
        definitions[(definition.start_date, definition.end_date)] = definition

    if not definitions:
        raise ValueError("El archivo no contiene fechas importables.")

    cycles: dict[tuple[date, date], PayrollCycle] = {}
    created = 0
    for key, definition in definitions.items():
        cycle = db.scalar(
            select(PayrollCycle).where(
                PayrollCycle.start_date == definition.start_date,
                PayrollCycle.end_date == definition.end_date,
            )
        )
        if cycle is None:
            cycle = PayrollCycle(
                cycle_name=definition.cycle_name,
                start_date=definition.start_date,
                end_date=definition.end_date,
            )
            db.add(cycle)
            db.flush()
            created += 1
        cycles[key] = cycle
    return cycles, created


def base_values(
    *,
    cycle_id: int,
    source_type: str,
    role_type: str,
    person_name: Any,
    person_code: Any,
    person_slot: str,
    row_number: int,
    source_hash: str,
    work_date: date,
    duration_minutes: int | None,
    status: Any,
) -> dict[str, Any]:
    values: dict[str, Any] = {field_name: ZERO for field_name in COMMON_ZERO_FIELDS}
    values.update(
        {
            "cycle_id": cycle_id,
            "source_type": source_type,
            "cost_center": source_type,
            "role_type": role_type,
            "source_employee_name": normalize_text(person_name),
            "source_employee_code": normalize_text(person_code) or None,
            "source_row_number": row_number,
            "source_row_hash": source_hash,
            "source_person_slot": person_slot,
            "work_date": work_date,
            "duration_minutes": duration_minutes,
            "status": normalize_text(status) or None,
        }
    )
    return values


def parse_workbook(
    content: bytes,
    source_type: str,
    cycles: dict[tuple[date, date], PayrollCycle],
) -> ParsedImport:
    workbook, sheet, expected_headers = load_source_sheet(content, source_type)

    result = ParsedImport(source_type=source_type, cost_center=source_type)
    relevant_indexes = [
        index for index, header in enumerate(expected_headers)
        if header not in {"Fecha Final", "URL", None}
    ]

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(row[: len(expected_headers)])
        if not any(value not in (None, "") for value in values):
            continue
        result.rows_read += 1
        date_index = 3 if source_type == "DR" else 4
        duration_index = 5 if source_type == "DR" else 6
        status_index = 6 if source_type == "DR" else 7
        work_date = to_work_date(values[date_index], workbook.epoch)
        if work_date is None:
            raise ValueError(f"Fila {row_number}: Fecha Inicial inválida.")
        definition = cycle_definition_for_date(work_date)
        cycle = cycles[(definition.start_date, definition.end_date)]
        if not is_valid_person(values[0]):
            result.errors.append(f"Fila {row_number}: Operador inválido.")
            continue

        source_hash = row_hash(values, relevant_indexes)
        duration = to_duration_minutes(values[duration_index])
        concepts = DR_CONCEPTS if source_type == "DR" else SERVICES_CONCEPTS

        driver = base_values(
            cycle_id=cycle.id,
            source_type=source_type,
            role_type="DRIVER",
            person_name=values[0],
            person_code=values[2] if source_type == "DR" else values[1],
            person_slot="OPERATOR",
            row_number=row_number,
            source_hash=source_hash,
            work_date=work_date,
            duration_minutes=duration,
            status=values[status_index],
        )
        driver.update({field_name: to_decimal(values[index]) for index, field_name in concepts.items()})
        result.candidates.append(RecordCandidate(driver))

        assistant_slots = [(1, "AUXILIARY")] if source_type == "DR" else [
            (2, "AUXILIARY_1"),
            (3, "AUXILIARY_2"),
        ]
        for index, slot in assistant_slots:
            if not is_valid_person(values[index]):
                continue
            assistant = base_values(
                cycle_id=cycle.id,
                source_type=source_type,
                role_type="ASSISTANT",
                person_name=values[index],
                person_code=None,
                person_slot=slot,
                row_number=row_number,
                source_hash=source_hash,
                work_date=work_date,
                duration_minutes=duration,
                status=values[status_index],
            )
            assistant.update(
                {field_name: to_decimal(values[column]) for column, field_name in concepts.items()}
            )
            if source_type == "DR":
                assistant["saturday_week_2_qty"] = to_decimal(values[26])
                assistant["sunday_week_2_qty"] = to_decimal(values[27])
            result.candidates.append(RecordCandidate(assistant))
    return result


def find_possible_reimports(
    db: Session,
    source_type: str,
    candidates: list[RecordCandidate],
) -> list[RecordCandidate]:
    existing: set[tuple[int, str, str, str]] = set()
    candidates_by_cycle: dict[int, list[RecordCandidate]] = {}
    for candidate in candidates:
        candidates_by_cycle.setdefault(candidate.values["cycle_id"], []).append(candidate)
    for cycle_id, cycle_candidates in candidates_by_cycle.items():
        hashes = sorted({item.values["source_row_hash"] for item in cycle_candidates})
        for start in range(0, len(hashes), 500):
            batch = hashes[start : start + 500]
            rows = db.execute(
                select(
                    PayrollRecord.cycle_id,
                    PayrollRecord.role_type,
                    PayrollRecord.source_person_slot,
                    PayrollRecord.source_row_hash,
                ).where(
                    PayrollRecord.cycle_id == cycle_id,
                    PayrollRecord.source_type == source_type,
                    PayrollRecord.cost_center == source_type,
                    PayrollRecord.source_row_hash.in_(batch),
                )
            )
            existing.update(tuple(row) for row in rows)
    return [candidate for candidate in candidates if candidate.duplicate_key in existing]


def persist_import(
    db: Session,
    parsed: ParsedImport,
    user: User,
    file_name: str,
) -> tuple[list[PayrollImport], int]:
    candidates_by_cycle: dict[int, list[RecordCandidate]] = {}
    for candidate in parsed.candidates:
        candidates_by_cycle.setdefault(candidate.values["cycle_id"], []).append(candidate)

    imports_by_cycle: dict[int, PayrollImport] = {}
    for cycle_id, candidates in candidates_by_cycle.items():
        payroll_import = PayrollImport(
            cycle_id=cycle_id,
            source_type=parsed.source_type,
            cost_center=parsed.cost_center,
            file_name=file_name,
            imported_by=user.id,
            rows_imported=len(candidates),
        )
        db.add(payroll_import)
        db.flush()
        imports_by_cycle[cycle_id] = payroll_import

    employees = list(db.scalars(select(Employee).order_by(Employee.id)).all())
    contract_map: dict[str, str | None] = {}
    for employee in employees:
        contract_map.setdefault(employee.employee_name.casefold(), employee.contract_type)

    created_employees = 0
    for candidate in parsed.candidates:
        values = candidate.values
        employee = next(
            (
                item
                for item in employees
                if item.role_type == values["role_type"]
                and names_refer_to_same_person(item.employee_name, values["source_employee_name"])
            ),
            None,
        )
        if employee is None:
            employee = Employee(
                employee_name=values["source_employee_name"],
                role_type=values["role_type"],
                contract_type=contract_map.get(values["source_employee_name"].casefold()),
            )
            db.add(employee)
            db.flush()
            employees.append(employee)
            contract_map.setdefault(values["source_employee_name"].casefold(), employee.contract_type)
            created_employees += 1
        db.add(
            PayrollRecord(
                **values,
                import_id=imports_by_cycle[values["cycle_id"]].id,
                employee_id=employee.id,
            )
        )
    db.commit()
    payroll_imports = list(imports_by_cycle.values())
    for payroll_import in payroll_imports:
        db.refresh(payroll_import)
    return payroll_imports, created_employees
