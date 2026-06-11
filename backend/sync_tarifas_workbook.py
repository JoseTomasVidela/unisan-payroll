from __future__ import annotations

import argparse
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from unicodedata import normalize

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import build_engine, is_sqlite_url
from app.models import Employee, PayrollConcept, PayrollConceptRate, PayrollCycle
from seed_payroll_concepts import apply_base_concepts

SOURCE_FILE = Path(
    r"C:\Users\jtvid\OneDrive\Escritorio\Acsa\Unisan\Planillas liquidaciones\Tarifas.xlsx"
)

ACTIVITY_TO_CONCEPT = {
    "DESPACHO / RETIRO": "DISPATCH_RETRIEVAL",
    "ENTRADA < 19:30": "ENTRY_BEFORE_1930",
    "ENTRADA < 07:30": "ENTRY_BEFORE_0730",
    "SALIDA > 19:30": "EXIT_AFTER_1930",
    "FERIA SEMANA 01": "FAIR_WEEK_1",
    "FERIA SEMANA 02": "FAIR_WEEK_2",
    "FUERA RADIO NORMAL": "OUTSIDE_RADIUS",
    "FUERA RADIO V REGION": "OUTSIDE_RADIUS_V_REGION",
    "SABADO SEMANA 01": "SATURDAY_WEEK_1",
    "DOMINGO SEMANA 01": "SUNDAY_WEEK_1",
    "VIAJES POR CLIENTE": "CLIENT_TRIPS",
    "SABADO > 16:00": "SATURDAY_AFTER_1600",
    "DOMINGO > 16:00": "SUNDAY_AFTER_1600",
    "SABADO SEMANA 02": "SATURDAY_WEEK_2",
    "DOMINGO SEMANA 02": "SUNDAY_WEEK_2",
    "SECADO FIN DE SEMANA": "WEEKEND_DRYING",
    "EVENTO": "EVENT",
    "PUNTO DE AGUA": "WATER_POINT",
    "BASURERO GRANDE": "LARGE_TRASH_BIN",
    "BASURERO CHICO": "SMALL_TRASH_BIN",
    "FOSA": "FOSA",
    "ASEO": "CLEANING",
    "SECADO": "DRYING",
    "ENTREGA KIT": "KIT_DELIVERY",
    "CARGA LAVAMANOS": "LAVATORY_LOAD",
    "SABADO": "SATURDAY_WEEK_1",
    "DOMINGO": "SUNDAY_WEEK_1",
    "ASEO FIN DE SEMANA": "WEEKEND_CLEANING",
    "SUCCION RILES (M3)": "RILES_SUCTION",
}


@dataclass(frozen=True)
class WorkbookRate:
    cost_center: str
    role_type: str
    contract_type: str
    concept_code: str
    amount: Decimal


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.replace("\xa0", " ").strip().split()).upper()


def normalize_cost_center(value: object) -> str:
    text = normalize_text(value)
    if text == "D&R":
        return "DR"
    if text == "SERVICIOS":
        return "SERVICES"
    raise ValueError(f"Centro de costo no reconocido: {value!r}")


def normalize_role_type(value: object) -> str:
    text = normalize_text(value)
    if "CONDUCTOR" in text:
        return "DRIVER"
    if "AUXILIAR" in text:
        return "ASSISTANT"
    raise ValueError(f"Cargo no reconocido: {value!r}")


def normalize_contract_type(value: object) -> str:
    text = normalize_text(value)
    if text == "ANTIGUO":
        return "OLD"
    if text == "NUEVO":
        return "NEW"
    raise ValueError(f"Contrato no reconocido: {value!r}")


def normalize_activity(value: object) -> str:
    text = normalize_text(value)
    prefixes = (
        "AUX ",
        "SERVICIO AUX ",
        "SERVICIO CHOFER ",
        "SERVICIOCHOFER ",
    )
    for prefix in prefixes:
        if text.startswith(prefix):
            text = text[len(prefix) :]
            break
    return text


def read_workbook_rates() -> list[WorkbookRate]:
    workbook = load_workbook(SOURCE_FILE, data_only=True, read_only=True)
    if "Tarifas" not in workbook.sheetnames:
        raise SystemExit("La hoja obligatoria 'Tarifas' no existe.")
    sheet = workbook["Tarifas"]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    expected = ("Centro Costo", "Cargo", "Actividad", "Precio", "Contrato")
    if header != expected:
        raise SystemExit(f"Encabezado inesperado en hoja Tarifas: {header!r}")
    result: list[WorkbookRate] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        center, cargo, activity, price, contract = row
        if not any(value not in (None, "") for value in row):
            continue
        normalized_activity = normalize_activity(activity)
        concept_code = ACTIVITY_TO_CONCEPT.get(normalized_activity)
        if concept_code is None:
            raise SystemExit(f"Actividad sin mapping: {activity!r}")
        result.append(
            WorkbookRate(
                cost_center=normalize_cost_center(center),
                role_type=normalize_role_type(cargo),
                contract_type=normalize_contract_type(contract),
                concept_code=concept_code,
                amount=Decimal(str(price)).quantize(Decimal("0.0001")),
            )
        )
    return result


def sync_workers_contracts(db: Session) -> None:
    for employee in db.scalars(select(Employee)).all():
        if employee.contract_type is None:
            employee.contract_type = None


def apply_rates(db: Session) -> None:
    apply_base_concepts(db)
    earliest_cycle = db.scalar(select(PayrollCycle).order_by(PayrollCycle.start_date).limit(1))
    effective_from_cycle_id = earliest_cycle.id if earliest_cycle else None
    rates = read_workbook_rates()
    for item in rates:
        concept = db.scalar(
            select(PayrollConcept).where(
                PayrollConcept.concept_code == item.concept_code,
                PayrollConcept.cost_center == item.cost_center,
                PayrollConcept.role_type == item.role_type,
            )
        )
        if concept is None:
            raise SystemExit(
                f"No existe concepto base para {item.cost_center}/{item.role_type}/{item.concept_code}"
            )
        existing = list(
            db.scalars(
                select(PayrollConceptRate).where(
                    PayrollConceptRate.concept_id == concept.id,
                    PayrollConceptRate.contract_type == item.contract_type,
                    PayrollConceptRate.active.is_(True),
                )
            ).all()
        )
        if len(existing) > 1:
            raise SystemExit(
                f"Mas de una tarifa activa para {item.cost_center}/{item.role_type}/{item.contract_type}/{item.concept_code}"
            )
        rate = existing[0] if existing else PayrollConceptRate(concept_id=concept.id, active=True)
        if not existing:
            db.add(rate)
        rate.amount = item.amount
        rate.contract_type = item.contract_type
        if rate.effective_from_cycle_id is None:
            rate.effective_from_cycle_id = effective_from_cycle_id
    db.commit()


def preview() -> None:
    rates = read_workbook_rates()
    print(f"Fuente: {SOURCE_FILE}")
    print(f"Tarifas detectadas: {len(rates)}")
    for item in rates[:20]:
        print(
            f"{item.cost_center:8} {item.role_type:9} {item.contract_type:3} "
            f"{item.concept_code:24} {item.amount}"
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Sincroniza tarifas desde Tarifas.xlsx")
    parser.add_argument("--apply", action="store_true", help="Aplicar en SQLite local")
    args = parser.parse_args()
    preview()
    if not args.apply:
        return
    settings = get_settings()
    if not is_sqlite_url(settings.database_url):
        raise SystemExit("Este sync esta bloqueado para MySQL/Azure.")
    engine = build_engine(settings.database_url)
    with Session(engine) as db:
        apply_rates(db)
    print("Tarifas sincronizadas en SQLite local.")


if __name__ == "__main__":
    main()
