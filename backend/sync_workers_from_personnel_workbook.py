from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import inspect, select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import engine
from app.employee_names import names_refer_to_same_person, parse_personnel_name
from app.main import bootstrap, normalize_cargo, normalize_email, normalize_rut
from app.models import Employee


@dataclass
class WorkbookWorker:
    raw_name: str
    display_name: str
    full_name: str
    first_name: str | None
    middle_name: str | None
    paternal_surname: str | None
    maternal_surname: str | None
    rut: str | None
    email: str | None
    cargo: str | None


def format_rut(number_value, digit_value) -> str | None:
    if number_value in (None, ""):
        return None
    rut_number = "".join(character for character in str(number_value) if character.isdigit())
    verifier = str(digit_value).strip().upper() if digit_value not in (None, "") else ""
    if not rut_number or not verifier:
        return None
    return f"{rut_number}-{verifier}"


def ensure_employee_cargo_column() -> None:
    inspector = inspect(engine)
    columns = {column["name"] for column in inspector.get_columns("payroll_employees")}
    statements: list[str] = []
    if "cargo" not in columns:
        statements.append("ALTER TABLE payroll_employees ADD COLUMN cargo VARCHAR(180) NULL")
    if "first_name" not in columns:
        statements.append("ALTER TABLE payroll_employees ADD COLUMN first_name VARCHAR(80) NULL")
    if "middle_name" not in columns:
        statements.append("ALTER TABLE payroll_employees ADD COLUMN middle_name VARCHAR(80) NULL")
    if "paternal_surname" not in columns:
        statements.append("ALTER TABLE payroll_employees ADD COLUMN paternal_surname VARCHAR(80) NULL")
    if "maternal_surname" not in columns:
        statements.append("ALTER TABLE payroll_employees ADD COLUMN maternal_surname VARCHAR(80) NULL")
    if not statements:
        return
    with engine.begin() as connection:
        for statement in statements:
            connection.exec_driver_sql(statement)


def load_workbook_workers(workbook_path: Path) -> list[WorkbookWorker]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "LISTADO" not in workbook.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'LISTADO'.")
    sheet = workbook["LISTADO"]
    header = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    expected = ["RUT", "dig", "NOMBRES", "CARGO", "CORREO"]
    if header[: len(expected)] != expected:
        raise ValueError(f"Encabezados inválidos. Esperados: {expected}. Recibidos: {header[:len(expected)]}")

    workers: list[WorkbookWorker] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        parsed_name = parse_personnel_name(row[2])
        workers.append(
            WorkbookWorker(
                raw_name=parsed_name.raw_name,
                display_name=parsed_name.display_name,
                full_name=parsed_name.full_name,
                first_name=parsed_name.given_names[0] if len(parsed_name.given_names) >= 1 else None,
                middle_name=parsed_name.given_names[1] if len(parsed_name.given_names) >= 2 else None,
                paternal_surname=parsed_name.paternal_surname or None,
                maternal_surname=parsed_name.maternal_surname or None,
                rut=normalize_rut(format_rut(row[0], row[1])),
                email=normalize_email(row[4]),
                cargo=normalize_cargo(str(row[3]) if row[3] is not None else None),
            )
        )
    return workers


def sync_workers(workbook_path: Path) -> dict[str, int]:
    bootstrap(get_settings(), engine)
    ensure_employee_cargo_column()
    workbook_workers = load_workbook_workers(workbook_path)

    created = 0
    updated = 0
    unchanged = 0

    with Session(engine) as db:
        employees = list(db.scalars(select(Employee).order_by(Employee.id)).all())
        for workbook_worker in workbook_workers:
            matches = [
                employee
                for employee in employees
                if any(
                    names_refer_to_same_person(employee.employee_name, candidate)
                    for candidate in (
                        workbook_worker.raw_name,
                        workbook_worker.display_name,
                        workbook_worker.full_name,
                    )
                )
            ]
            if matches:
                changed = False
                for employee in matches:
                    if not employee.rut and workbook_worker.rut:
                        employee.rut = workbook_worker.rut
                        changed = True
                    if not employee.email and workbook_worker.email:
                        employee.email = workbook_worker.email
                        changed = True
                    if not employee.cargo and workbook_worker.cargo:
                        employee.cargo = workbook_worker.cargo
                        changed = True
                    if employee.first_name != workbook_worker.first_name:
                        employee.first_name = workbook_worker.first_name
                        changed = True
                    if employee.middle_name != workbook_worker.middle_name:
                        employee.middle_name = workbook_worker.middle_name
                        changed = True
                    if employee.paternal_surname != workbook_worker.paternal_surname:
                        employee.paternal_surname = workbook_worker.paternal_surname
                        changed = True
                    if employee.maternal_surname != workbook_worker.maternal_surname:
                        employee.maternal_surname = workbook_worker.maternal_surname
                        changed = True
                updated += 1 if changed else 0
                unchanged += 0 if changed else 1
                continue

            employee = Employee(
                employee_name=workbook_worker.display_name or workbook_worker.full_name or workbook_worker.raw_name,
                role_type="UNASSIGNED",
                rut=workbook_worker.rut,
                email=workbook_worker.email,
                cargo=workbook_worker.cargo,
                first_name=workbook_worker.first_name,
                middle_name=workbook_worker.middle_name,
                paternal_surname=workbook_worker.paternal_surname,
                maternal_surname=workbook_worker.maternal_surname,
            )
            db.add(employee)
            db.flush()
            employees.append(employee)
            created += 1

        db.commit()

    return {
        "rows_read": len(workbook_workers),
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Sincroniza trabajadores desde LISTADO PERSONAL.")
    parser.add_argument("workbook_path", help="Ruta al archivo Excel origen.")
    args = parser.parse_args()
    workbook_path = Path(args.workbook_path)
    if not workbook_path.exists():
        raise SystemExit(f"No se encontró el archivo: {workbook_path}")
    result = sync_workers(workbook_path)
    print(result)


if __name__ == "__main__":
    os.environ.setdefault("PAYROLL_DATABASE_URL", "sqlite:///./payroll_dev.db")
    os.environ.setdefault("PAYROLL_JWT_SECRET", "un-secreto-local-de-al-menos-32-caracteres")
    main()
