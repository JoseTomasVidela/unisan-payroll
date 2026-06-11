from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import unicodedata

from openpyxl import load_workbook

RATE_SCALE = Decimal("0.0001")


@dataclass(frozen=True)
class RateOccurrence:
    sheet_name: str
    cell: str
    row_number: int
    raw_label: str
    amount: Decimal


def normalize_label(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return " ".join(ascii_value.casefold().split())


def source_labels(source_type: str, role_type: str, concept_name: str) -> set[str]:
    prefixes = [""]
    if source_type == "DR" and role_type == "ASSISTANT":
        prefixes = ["Aux "]
    elif source_type == "SERVICES" and role_type == "DRIVER":
        prefixes = ["Servicio Chofer ", "ServicioChofer "]
    elif source_type == "SERVICES" and role_type == "ASSISTANT":
        prefixes = ["Servicio Aux "]
    return {normalize_label(prefix + concept_name) for prefix in prefixes}


def extract_rate_occurrences(
    source_file: Path,
    seed_items,
) -> dict[tuple[str, str, str], list[RateOccurrence]]:
    label_to_key: dict[str, tuple[str, str, str]] = {}
    for seed in seed_items:
        key = (seed.source_type, seed.role_type, seed.concept_code)
        for label in source_labels(seed.source_type, seed.role_type, seed.concept_name):
            if label in label_to_key and label_to_key[label] != key:
                raise ValueError(f"Etiqueta fuente ambigua: {label}")
            label_to_key[label] = key

    workbook = load_workbook(source_file, read_only=True, data_only=True)
    occurrences: dict[tuple[str, str, str], list[RateOccurrence]] = defaultdict(list)
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_col=1, max_col=3, values_only=True),
            start=1,
        ):
            raw_label, _, raw_rate = row
            if not isinstance(raw_label, str) or not isinstance(raw_rate, (int, float, Decimal)):
                continue
            key = label_to_key.get(normalize_label(raw_label))
            if key is None:
                continue
            occurrences[key].append(
                RateOccurrence(
                    sheet_name=sheet.title,
                    cell=f"C{row_number}",
                    row_number=row_number,
                    raw_label=raw_label.strip(),
                    amount=Decimal(str(raw_rate)).quantize(RATE_SCALE, rounding=ROUND_HALF_UP),
                )
            )
    return dict(occurrences)


def rate_options(occurrences: list[RateOccurrence]) -> dict[Decimal, list[RateOccurrence]]:
    options: dict[Decimal, list[RateOccurrence]] = defaultdict(list)
    for occurrence in occurrences:
        options[occurrence.amount].append(occurrence)
    return dict(options)
